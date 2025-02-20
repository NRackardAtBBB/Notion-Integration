from notion_client import Client
import os
import openpyxl
from typing import Optional
import uuid
from src.utils.notion_response_cleaner import clean_text_chunk_record
import logging
from config.notion_config import NotionConfig
from src.models.models import Project, TextChunk

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

class NotionWrapper:
    def __init__(self, test_mode: bool = False):
        self.test_mode = test_mode
        self.config = NotionConfig()
        self.client = Client(auth=self.config.api_key)
        self.test_file = os.path.join("data", "test_notion_pages.xlsx")
        self.project_table_id, self.text_table_id = self.initialize_notion_databases()
    
    # Database Operations
    def initialize_notion_databases(self):
            """Initialize or get Notion databases for projects and text library"""
            projects_id = self.find_database_by_name("BBB Projects")
            text_library_id = self.find_database_by_name("BBB Text Library")
            
            if not projects_id:
                projects_id = self.create_database(
                    self.config.page_id,
                    self.config.DBTables.PROJECTS.value,
                    Project.database_schema()
                )

            if not text_library_id:
                text_library_id = self.create_database(
                    self.config.page_id, 
                    self.config.DBTables.TEXT_LIBRARY.value, 
                    TextChunk.database_schema(projects_id)
                )
                
            return projects_id, text_library_id

    def create_database(self, parent_page_id: str, db_name: str, properties: dict) -> str:
        """
        Create a new database under a given page and return its ID.
        """
        new_database = {
            "parent": {"page_id": parent_page_id},
            "title": [ {"type": "text", "text": {"content": db_name } } ],
            "properties": properties
        }

        response = self.client.databases.create(**new_database)
        return response["id"]

    def update_database(self, database_id: str, properties: dict) -> None:
        """Updates existing database properties"""
        return self.client.databases.update(database_id=database_id, properties=properties)
    
    def find_database_by_name(self, query: str) -> Optional[str]:
        """Searches for database by name and returns its ID if found"""

        results = self.client.search(
            query=query,
            filter={"property": "object", "value": "database"}
        ).get("results")
        
        return results[0]["id"] if results else None
    

    # Page Operations
    def create_page(self, database_id: str, properties: dict) -> dict:
        """Creates a new page in specified database"""
        new_page = {
            "parent": {"database_id": database_id},
            "properties": properties
        }
        if self.test_mode:
            dummy_response = new_page.copy()
            dummy_response["id"] = "test-id-" + str(uuid.uuid4())
            self._save_to_excel(dummy_response)
            return dummy_response
        else:
            return self.client.pages.create(**new_page)
        

    # Project and Text Management
    def add_projects(self, projects):
        """Add projects to Notion and return mapping of project numbers to Notion IDs"""
        project_ids = {}
        for project in projects:
            response = self.create_page(self.project_table_id, project.to_notion_properties())
            project_ids[project.project_number] = response["id"]
            logging.info(f"Added project: {project.name}")
        return project_ids
    
    def add_text_chunks(self, text_chunks, project_ids):
        """Add text chunks to Notion with project relations"""
        for chunk in text_chunks:
            notion_uuid = project_ids[chunk.project_number]
            chunk.project_id = notion_uuid
            self.create_page(self.text_table_id, chunk.to_notion_properties())
            logging.info(f"Added chunk: {chunk.title}")

    def _save_to_excel(self, page_data: dict) -> None:
        """
        Save the page data to an Excel file.
        We'll first clean the properties (e.g. flatten nested rich_text objects) 
        using our helper in notion_response_cleaner, and then use the database id 
        (from page_data['parent']['database_id']) as the sheet name so that pages 
        belonging to different tables end up in different sheets.
        """
        # If the page data has a 'properties' field, clean the properties for readability.
        if "properties" in page_data:
            cleaned_properties = clean_text_chunk_record(page_data)
            # Optionally merge with parent data if desired; here we override page_data 
            # with our cleaned version.
            page_data = cleaned_properties

        # Determine the sheet name based on the database id
        db_id = page_data.get("parent.database_id", "UnknownTable")
        sheet_name = f"Table_{db_id[-4:]}"  # using the last 4 characters for uniqueness
        
        # Load existing workbook or create a new one
        if os.path.exists(self.test_file):
            wb = openpyxl.load_workbook(self.test_file)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                default_sheet = wb["Sheet"]
                wb.remove(default_sheet)
        
        # Check if sheet exists; if not, create it with headers based on flattened keys
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            headers = list(self._flatten_page_data(page_data).keys())
            ws.append(headers)
        
        # Flatten the page_data then match the header order
        flat_data = self._flatten_page_data(page_data)
        header = [cell.value for cell in ws[1]]
        row = [str(flat_data.get(key, "")) for key in header]
        ws.append(row)
        wb.save(self.test_file)
    
    def _flatten_page_data(self, page_data: dict, parent_key: str = "", sep: str = ".") -> dict:
        """
        Flatten a nested dictionary. For example, turns:
          {"parent": {"database_id": "123"}, "properties": {"Title": "ABC"}}
        into:
          {"parent.database_id": "123", "properties.Title": "ABC"}
        """
        items = {}
        for k, v in page_data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(self._flatten_page_data(v, new_key, sep=sep))
            else:
                items[new_key] = v
        return items
    
